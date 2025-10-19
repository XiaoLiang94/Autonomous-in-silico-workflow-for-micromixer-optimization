# -*- coding: utf-8 -*-
"""
@author: Xiao Liang
"""

import subprocess
import os
import time
import win32com.client as win32
import shutil
import pandas as pd
import numpy as np
import re
import random
import networkx as nx
from scipy.spatial import distance
from openpyxl import Workbook, load_workbook
import comtypes.client
import pythoncom
import matplotlib
matplotlib.use('Qt5Agg')  # Use the TkAgg backend
import matplotlib.pyplot as plt


## Functions for SolidWorks control
# ----------------------------------------------------------------------------------------------------------------------------
 
def open_sldprt_and_run_macro(file_path, macro_path, macro_module1, macro_procedure1, macro_module2, macro_procedure2):
    """
    Use cmd to call SolidWorks, open a specified .SLDPRT file, and run two specified macros.
    :param file_path: Path to the .SLDPRT file
    :param macro_path: Path to the macro file
    :param macro_module1: Module name of macro 1
    :param macro_procedure1: Procedure name of macro 1
    :param macro_module2: Module name of macro 2
    :param macro_procedure2: Procedure name of macro 2
    """
    # Check if file path exists
    if not os.path.exists(file_path):
        print(f"File path does not exist: {file_path}")
        return

    # SolidWorks executable path
    solidworks_exe = r"D:\XXXXX\XXXXX\SLDWORKS.exe"
    if not os.path.exists(solidworks_exe):
        print(f"SolidWorks executable path does not exist: {solidworks_exe}")
        return

    # Initialize COM environment
    pythoncom.CoInitialize()

    try:
        # Command to open SolidWorks
        command = f'"{solidworks_exe}" "{file_path}"'
        print(f"Executing command: {command}")
        process = subprocess.Popen(command, shell=True)
        print("Successfully opened the SolidWorks file!")

        # Wait until SolidWorks starts
        time.sleep(60)

        # Initialize SolidWorks application
        swApp = win32.Dispatch("SldWorks.Application")
        swApp.Visible = True
        print("Successfully connected to SolidWorks application.")

        # Verify macro file existence
        if not os.path.exists(macro_path):
            print(f"Macro file path does not exist: {macro_path}")
            return

        # Run the first macro
        print(f"Running macro 1: Module={macro_module1}, Procedure={macro_procedure1}")
        result1 = swApp.RunMacro(macro_path, macro_module1, macro_procedure1)
        if result1 == True:
            print("Macro 1 executed successfully!")
        else:
            print(f"Macro 1 execution failed, error code: {result1}")

        # Run the second macro
        print(f"Running macro 2: Module={macro_module2}, Procedure={macro_procedure2}")
        result2 = swApp.RunMacro(macro_path, macro_module2, macro_procedure2)
        if result2 == True:
            print("Macro 2 executed successfully!")
        else:
            print(f"Macro 2 execution failed, error code: {result2}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Ensure SolidWorks is properly closed
        if swApp:
            try:
                print("Closing all opened documents...")
                swApp.CloseAllDocuments()
                print("All documents have been closed.")

                print("Closing SolidWorks...")
                swApp.Quit()
                print("SolidWorks has been closed.")
            except Exception as e:
                print(f"Error closing SolidWorks: {e}")

        # Ensure processes are terminated
        try:
            print("Checking and closing SolidWorks-related processes...")
            os.system("taskkill /F /IM SLDWORKS.exe")
            print("SolidWorks processes have been terminated.")
        except Exception as e:
            print(f"Unable to terminate SolidWorks processes: {e}")

        # Release COM objects
        try:
            print("Releasing SolidWorks COM objects...")
            comtypes.CoUninitialize()
            print("SolidWorks COM objects have been released.")
            
            # Release COM environment
            pythoncom.CoUninitialize()
            
        except Exception as e:
            print(f"Error releasing COM objects: {e}")


def update_bas_file(original_file_path, changes, i):
    """ 
    Updates a .bas file by replacing specified values and returns the modified content.

    Parameters:
        original_file_path (str): Path to the original .bas file.
        changes (dict): Dictionary of replacements in the format {old_value: new_value}.
        i (int): Current iteration index to replace placeholder values.

    Returns:
        list: Modified lines of the .bas file.
    """
    try:
        # Step 1: Read the original .bas file
        with open(original_file_path, 'r') as file:
            lines = file.readlines()

        # Step 2: Modify the content based on the iteration
        for line_index, line in enumerate(lines):
            for old_value, new_value in changes.items():
                if old_value in line:
                    lines[line_index] = line.replace(old_value, new_value.replace("{i}", str(i)))

        return lines

    except Exception as e:
        print(f"An error occurred: {e}")
        return []       
    
    
def copy_and_rename_macro_file(src_file, dest_dir, i):
    """
    Copies and renames the macro file for the current iteration.

    Parameters:
        src_file (str): Source macro file path.
        dest_dir (str): Destination directory for the renamed file.
        i (int): Current iteration index.

    Returns:
        str: Full path of the copied and renamed file.
    """
    try:
        # Create the destination directory if it doesn't exist
        os.makedirs(dest_dir, exist_ok=True)

        # Construct the destination file path
        dest_file = os.path.join(dest_dir, f"test_T_{i}.swp")

        # Copy and rename the file
        shutil.copy(src_file, dest_file)
        print(f"Copied and renamed macro file to: {dest_file}")

        return dest_file
    except Exception as e:
        print(f"An error occurred while copying and renaming the file: {e}")
        return None


# Functions for StarCCM+ control
# ----------------------------------------------------------------------------------------------------------------------------

def replace_strings_and_update_population(input_file_path, output_file_path, old_string, new_string, folder_path):
    """
    Open a Java file, replace specified strings, and update 'population_number' with the number of .x_t files in a folder. Save the modified file.

    :param input_file_path: Original Java file path
    :param output_file_path: Path to save the modified file (including file name)
    :param old_string: String to be replaced
    :param new_string: Replacement string
    :param folder_path: Folder to count the number of .x_t files
    """
    if not os.path.exists(input_file_path):
        print(f"Input file does not exist: {input_file_path}")
        return

    if not os.path.exists(folder_path):
        print(f"Folder does not exist: {folder_path}")
        return

    # Count the number of .x_t files
    population_number = len([f for f in os.listdir(folder_path) if f.endswith('.x_t')])

    output_dir = os.path.dirname(output_file_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")

    # Read the original file
    with open(input_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # Replace specified strings and population_number
    modified_lines = []
    for line in lines:
        line = line.replace(old_string, new_string)
        if "population_number" in line:
            # Replace population_number directly
            line = line.replace("population_number", str(population_number))
        modified_lines.append(line)

    # Save as a new file
    with open(output_file_path, 'w', encoding='utf-8') as file:
        file.writelines(modified_lines)

    print(f"Modified file saved to: {output_file_path}")


def run_starccm(batch_file_path):
    """
    Execute a starccm+ command using CMD and display the output in real-time.

    :param batch_file_path: Path to the modified Java file
    """
    try:
        # Locate the directory
        starccm_dir = r"C:\Program Files\Siemens\17.04.008\STAR-CCM+17.04.008\star\bin"
        
        # Construct the full CMD command
        commands = f'cd /d "{starccm_dir}" && starccm+ -np 16 -batch "{batch_file_path}"'

        # Start the subprocess
        process = subprocess.Popen(
            commands,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,  # Recommended in Python 3.7+
        )

        # Read CMD output in real-time
        for line in process.stdout:
            print(line, end="")  # Print standard output
        for line in process.stderr:
            print(line, end="")  # Print error output

        process.wait()  # Wait for the subprocess to complete
        print("starccm+ command has finished running.")
    except Exception as e:
        print(f"An error occurred while running starccm+ command: {e}")

        
        
## Functions for editing excel     
# ----------------------------------------------------------------------------------------------------------------------------

def natural_key(string):
    """Key function for natural sorting, extracting numeric parts of a string."""
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', string)]


def process_csv(file_path, output_folder):
    """
    Processes a CSV file, calculates MI values for each plate, and saves results to an Excel file.

    :param file_path: Path to the CSV file
    :param output_folder: Directory to save the processed Excel files
    :return: A tuple containing the base file name, MI values, and the F2 value
    """
    df = pd.read_csv(file_path)

    # Filter the data into separate DataFrames based on the X values
    plates = {
        'plate1': df[df['X (m)'] <= 0.001],
        'plate2': df[df['X (m)'] == 0.002],
        'plate3': df[df['X (m)'] == 0.003],
        'plate4': df[df['X (m)'] == 0.004],
        'plate5': df[df['X (m)'] == 0.005]
    }

    # Generate output Excel file name based on the input CSV file name
    base_name = os.path.basename(file_path)
    output_file_name = os.path.splitext(base_name)[0] + '_restructured.xlsx'
    output_path = os.path.join(output_folder, output_file_name)

    mi_values = {}

    # Extract the value of the F2 cell (row 0, column 5)
    f2_value = df.iloc[0, 5] if df.shape[1] > 5 and df.shape[0] > 0 else None

    # Create a new Excel writer object
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        for plate_name, plate_data in plates.items():
            # Pivot the data
            pivot_table = plate_data.pivot_table(values='PS', index='Z (m)', columns='Y (m)', aggfunc='mean')

            # Calculate Average and (c-ci)^2 directly in Python
            pivot_table['Average'] = pivot_table.mean(axis=1)
            pivot_table['(c-ci)^2'] = (pivot_table['Average'] - 0.5) ** 2

            # Calculate MI value directly
            mi_value = 1 - (np.sqrt(pivot_table['(c-ci)^2'].mean()) / 0.5)
            mi_values[plate_name] = mi_value

            # Write the pivot table and MI value to a new sheet
            pivot_table.to_excel(writer, sheet_name=plate_name)
            worksheet = writer.sheets[plate_name]
            worksheet.write('AI1', f'MI_{plate_name}')
            worksheet.write('AI2', mi_value)

    print(f'File saved to: {output_path}')
    return base_name, mi_values, f2_value


def process_all_csv_files(input_folder, output_folder, summary_file):
    """
    Processes all CSV files in a folder and generates a summary file with MI values and averages.

    :param input_folder: Directory containing the input CSV files
    :param output_folder: Directory to save the processed files
    :param summary_file: Path to save the summary CSV file
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    summary_data = []

    # Get file names and sort them naturally
    filenames = [f for f in os.listdir(input_folder) if f.endswith(".csv")]
    filenames.sort(key=natural_key)

    for filename in filenames:
        file_path = os.path.join(input_folder, filename)
        base_name, mi_values, f2_value = process_csv(file_path, output_folder)
        summary_data.append((base_name, mi_values, f2_value))

    # Write the summary data to a CSV file
    summary_rows = []
    for base_name, mi_values, f2_value in summary_data:
        row = [base_name]
        row.extend(mi_values.values())
        average_value = np.mean(list(mi_values.values())) if mi_values else None
        row.append(average_value)  # Add the average value
        row.append(f2_value if f2_value is not None else 'N/A')
        summary_rows.append(row)

    # Create a DataFrame for summary data
    summary_df = pd.DataFrame(summary_rows, columns=['Design', 'plate1', 'plate2', 'plate3', 'plate4', 'plate5', 'obj1', 'obj2'])

    # Save summary to CSV
    summary_df.to_csv(summary_file, index=False)
    print(f'Summary file saved to: {summary_file}')



def read_summary_csv(summary_file):

    # Read the CSV file
    df = pd.read_csv(summary_file)
    
    # Assuming the last two columns are Mixing Index and Pressure Drop
    mixing_indices = df.iloc[:, -2].tolist()
    pressure_drops = df.iloc[:, -1].tolist()
    
    return mixing_indices, pressure_drops


## Functions for optimization algorithm   
# ----------------------------------------------------------------------------------------------------------------------------

class Mixer:
    """
    Custom problem for optimization.
    """
    def __init__(self):
        self.num_variables = 4
        self.num_objectives = 2
        self.lower_bounds = [1] * self.num_variables
        self.upper_bounds = [36] * self.num_variables
        self.edges = [
            (9, 13), (5, 9), (1, 5), (10, 14), (6, 10), (2, 6),
            (11, 15), (7, 11), (3, 7), (12, 16), (8, 12), (4, 8),
            (9, 10), (5, 6), (10, 11), (6, 7), (11, 12), (7, 8),
            (10, 13), (9, 14), (6, 9), (5, 10), (2, 5), (1, 6),
            (11, 14), (10, 15), (7, 10), (6, 11), (3, 6), (2, 7),
            (12, 15), (11, 16), (11, 8), (7, 12), (4, 7), (3, 8)
        ]

    def repair_solution(self, positions):
        """
        Repair solution to ensure no invalid connections, no duplicate edges,
        and maintain the required number of variables.
        """
        # Ensure positions are integers
        positions = [int(pos) for pos in positions]

        # Remove duplicates while preserving the order
        positions = list(dict.fromkeys(positions))

        # Get the edges corresponding to the positions
        selected_edges = [self.edges[pos - 1] for pos in positions if pos - 1 < len(self.edges)]

        # Build the graph from the selected edges
        g = nx.Graph()
        g.add_edges_from(selected_edges)

        # Explicitly add all possible nodes to the graph
        g.add_nodes_from(range(1, 17))  # Nodes are 1 to 16

        # Ensure no invalid top-to-bottom connections
        top_nodes = {1, 2, 3, 4}
        bottom_nodes = {13, 14, 15, 16}

        # Remove invalid connections
        while self.has_top_to_bottom_path(g, top_nodes, bottom_nodes):
            # Get the path causing the issue
            path = self.get_top_to_bottom_path(g, top_nodes, bottom_nodes)

            # Convert the path into edge tuples
            path_edges = [(path[i], path[i + 1]) for i in range(len(path) - 1)]

            # Find the first edge in the path to remove
            edge_to_remove = None
            for edge in path_edges:
                if edge in selected_edges or tuple(reversed(edge)) in selected_edges:
                    edge_to_remove = edge if edge in selected_edges else tuple(reversed(edge))
                    break

            if edge_to_remove:
                # Remove the edge from the graph and the selected edges
                selected_edges.remove(edge_to_remove)
                g.remove_edge(*edge_to_remove)

                # Add a new random edge not in the solution
                new_edge = random.choice(
                    [e for e in self.edges if e not in selected_edges and tuple(reversed(e)) not in selected_edges]
                )
                selected_edges.append(new_edge)
                g.add_edge(*new_edge)
            else:
                # If no edge can be removed, break to prevent infinite loops
                break

        # Ensure unique edges in the solution
        unique_edges = []
        edge_set = set()
        for edge in selected_edges:
            if edge not in edge_set and tuple(reversed(edge)) not in edge_set:
                unique_edges.append(edge)
                edge_set.add(edge)

        # Add new valid unique edges if the solution has fewer than the required number of variables
        while len(unique_edges) < self.num_variables:
            new_edge = random.choice(
                [e for e in self.edges if e not in edge_set and tuple(reversed(e)) not in edge_set]
            )
            unique_edges.append(new_edge)
            edge_set.add(new_edge)
            g.add_edge(*new_edge)

        # Map repaired edges back to their positions (ensure no duplicates in positions)
        repaired_positions = [self.edges.index(edge) + 1 for edge in unique_edges]
        return repaired_positions

    def has_top_to_bottom_path(self, g, top_nodes, bottom_nodes):
        """ Check if there is a path from any top node to any bottom node. """
        for top_node in top_nodes:
            for bottom_node in bottom_nodes:
                if nx.has_path(g, top_node, bottom_node):
                    return True
        return False

    def get_top_to_bottom_path(self, g, top_nodes, bottom_nodes):
        """ Get any path from top to bottom if it exists. """
        for top_node in top_nodes:
            for bottom_node in bottom_nodes:
                if nx.has_path(g, top_node, bottom_node):
                    return nx.shortest_path(g, top_node, bottom_node)
        return None

    def create_solution(self):
        """
        Create a random solution.
        """
        variables = random.sample(range(1, 37), self.num_variables)
        return {"variables": variables, "objectives": [0.0, 0.0]}


def calculate_hypervolume(front, reference_point):
    """
    Calculate the HyperVolume (HV) of the Pareto front for mixed objectives.

    Args:
        front (list of lists): The Pareto front as a list of [obj1, obj2] points.
        reference_point (list): The reference point as [ref_obj1, ref_obj2].

    Returns:
        float: The computed hypervolume.
    """
    if not front:
        print("no input front")
        return 0.0

    # Adjust the front for uniform minimization (invert the maximization objective)
    adjusted_front = [[-obj[0], obj[1]] for obj in front]

    # Sort the front by the first objective (obj1, descending)
    sorted_front = sorted(adjusted_front, key=lambda x: x[0], reverse=True)

    # Initialize hypervolume
    hv = 0.0

    # Calculate the hypervolume using rectangles
    for i in range(len(sorted_front)):
        current_point = sorted_front[i]

        # Calculate width (difference in obj1)
        if i == 0:
            width = reference_point[0] - current_point[0]
        else:
            width = sorted_front[i - 1][0] - current_point[0]

        # Calculate height (difference in obj2)
        height = reference_point[1] - current_point[1]

        # Only add positive areas
        if width > 0 and height > 0:
            hv += width * height

    return hv


def ensure_integer_variables(solution):
    """
    Ensure that all variables in a solution are integers.
    """
    solution["variables"] = [int(var) for var in solution["variables"]]
    return solution

def evaluate_offspring_from_file(offspring, filepath):
    """
    Load fitness values for offspring from a file and assign objectives.
    """
    data = pd.read_csv(filepath)
    print(f"Number of rows in file: {len(data)}")
    for i, solution in enumerate(offspring):
        solution["objectives"][0] = data.iloc[i, -2]  # Assuming second last column is Obj1
        solution["objectives"][1] = data.iloc[i, -1]  # Assuming last column is Obj2

    return offspring

def load_pre_existing_population(filepath, problem):
    """
    Load pre-existing solutions from the provided CSV file.
    """
    pre_existing_data = pd.read_csv(filepath)
    population = []
    for _, row in pre_existing_data.iterrows():
        solution = {
            "variables": [
                row["block1 position"],
                row["block2 position"],
                row["block3 position"],
                row["block4 position"],
            ],
            "objectives": [row["Obj1"], row["Obj2"]],
        }
        population.append(solution)
    return population

def identify_pareto_front(population, target_size=3):
    """
    Identify a population of size target_size, containing Pareto fronts in sequence.
    """
    pareto_fronts = []
    remaining_population = population[:]
    selected_population = []

    while remaining_population and len(selected_population) < target_size:
        current_front = []
        for i, candidate in enumerate(remaining_population):
            dominated = False
            for j, competitor in enumerate(remaining_population):
                if i != j and dominates(competitor, candidate):
                    dominated = True
                    break
            if not dominated:
                current_front.append(candidate)

        pareto_fronts.append(current_front)
        if len(selected_population) + len(current_front) <= target_size:
            selected_population.extend(current_front)
        else:
            selected_population.extend(current_front[: target_size - len(selected_population)])

        remaining_population = [ind for ind in remaining_population if ind not in current_front]

    return selected_population

def generate_initial_population(problem, population_size):
    """
    Generate an initial population with feasible solutions.
    """
    population = []
    while len(population) < population_size:
        solution = problem.create_solution()
        solution["variables"] = problem.repair_solution(solution["variables"])  # Ensure feasibility
        population.append(solution)
    return population

def dominates(solution_a, solution_b):
    """
    Check if solution_a dominates solution_b.
    """
    return dominates_solution(solution_a, solution_b)


def non_dominated_sorting(population):
    """
    Perform non-dominated sorting on the population.
    """
    fronts = []
    domination_counts = [0] * len(population)
    dominates = [set() for _ in range(len(population))]

    for i, sol_i in enumerate(population):
        for j, sol_j in enumerate(population):
            if dominates_solution(sol_i, sol_j):
                dominates[i].add(j)
            elif dominates_solution(sol_j, sol_i):
                domination_counts[i] += 1
        if domination_counts[i] == 0:
            sol_i["rank"] = 0
            if len(fronts) == 0:
                fronts.append([])
            fronts[0].append(i)

    current_rank = 0
    while len(fronts[current_rank]) > 0:
        next_front = []
        for i in fronts[current_rank]:
            for j in dominates[i]:
                domination_counts[j] -= 1
                if domination_counts[j] == 0:
                    population[j]["rank"] = current_rank + 1
                    next_front.append(j)
        fronts.append(next_front)
        current_rank += 1

    return [[population[i] for i in front] for front in fronts if len(front) > 0]


def dominates_solution(sol_a, sol_b):
    """
    Check if solution A dominates solution B.
    Maximize Objective 1 and Minimize Objective 2.
    """
    better_in_all = (sol_a["objectives"][0] >= sol_b["objectives"][0]) and (sol_a["objectives"][1] <= sol_b["objectives"][1])
    better_in_one = (sol_a["objectives"][0] > sol_b["objectives"][0]) or (sol_a["objectives"][1] < sol_b["objectives"][1])
    return better_in_all and better_in_one


def tournament_selection(population, k=2):
    """
    Perform tournament selection.
    """
    selected = random.sample(population, k)
    return min(selected, key=lambda sol: sol["rank"])

def crossover(parent1, parent2, crossover_rate, problem=None):
    """
    Perform single-point crossover and ensure unique and valid offspring.
    """
    if random.random() > crossover_rate:
        # If crossover doesn't occur, return parents as offspring
        # print(f"No crossover applied. Returning parents as offspring.")
        return parent1, parent2

    # Choose a random crossover point
    point = random.randint(1, len(parent1["variables"]) - 1)

    # Combine parts from both parents
    child1_vars = parent1["variables"][:point] + parent2["variables"][point:]
    child2_vars = parent2["variables"][:point] + parent1["variables"][point:]

    # print(f"Crossover point: {point}")
    # print(f"Before Repair - Child 1: {child1_vars}")
    # print(f"Before Repair - Child 2: {child2_vars}")

    # Create offspring solutions
    child1 = {"variables": child1_vars, "objectives": [0.0, 0.0]}
    child2 = {"variables": child2_vars, "objectives": [0.0, 0.0]}

    # Repair offspring to ensure validity
    if problem:
        child1["variables"] = problem.repair_solution(child1["variables"])
        child2["variables"] = problem.repair_solution(child2["variables"])

    # Ensure variables are integers
    child1 = ensure_integer_variables(child1)
    child2 = ensure_integer_variables(child2)

    return child1, child2

def is_duplicate(candidate, population):
    """
    Check if a candidate solution is already in the population.
    """
    candidate_vars = set(candidate["variables"])
    for solution in population:
        if set(solution["variables"]) == candidate_vars:
            return True
    return False

def mutate(solution, mutation_rate, problem=None):
    """
    Perform mutation on a solution and ensure unique and valid variables.
    """
    for i in range(len(solution["variables"])):
        if random.random() < mutation_rate:
            new_var = random.randint(1, 36)
            # Replace variable at index `i` with a new one that doesn't duplicate
            while new_var in solution["variables"]:
                new_var = random.randint(1, 36)
            solution["variables"][i] = new_var

    # Repair the solution to ensure no invalid connections
    if problem:
        solution["variables"] = problem.repair_solution(solution["variables"])
    solution["objectives"] = [0.0, 0.0]

    return ensure_integer_variables(solution)

def save_population_to_template(population, template_file, output_file, sheet_name, start_row, start_col):
    """
    Save the population to a specific Excel file, based on a template, always to 'simple' sheet.
    """
    # Load the template file
    try:
        workbook = load_workbook(template_file)
    except FileNotFoundError:
        raise FileNotFoundError(f"Template file '{template_file}' not found.")

    # Access or create the 'simple' sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Write population data to the target sheet
    for row_index, solution in enumerate(population, start=start_row):
        # Write the solution index
        sheet.cell(row=row_index, column=start_col, value=row_index - start_row + 1)  # Solution number
        # Write the variables
        for col_index, variable in enumerate(solution["variables"], start=start_col + 1):
            sheet.cell(row=row_index, column=col_index, value=variable)

    # Save the file as a new Excel file
    workbook.save(output_file)
    print(f"Population saved to '{output_file}' in the sheet '{sheet_name}'.")


     
## Mian
# ----------------------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------------

# File paths and related parameters
original_bas_file_path = r"D:\Close_loop_in_silico_optimization_showcase\Creating3D.bas"
output_directory = r"D:\Close_loop_in_silico_optimization_showcase"
file_path = r"D:\Close_loop_in_silico_optimization_showcase\Blank.SLDPRT"

# Original macro file path
src_macro_file = r"D:\Close_loop_in_silico_optimization_showcase\test.swp"

# Macro module and procedure names
macro_module_name = "Module1"
macro_procedure_name = "main"
macro_module_name2 = "Module2"
macro_procedure_name2 = "main"

# Problem definition
problem = Mixer()

# Algorithm parameters
population_size = 2
generations = 2
mutation_rate = 0.3
crossover_rate = 0.7

# Template file name
template_file = "Test.xlsx"

# Reference point for HyperVolume calculation
reference_point = [-1.0, 50.0]

# Generate initial population
initial_population = generate_initial_population(problem, population_size)

# Save Initial Population to Test_1.xlsx in 'simple' sheet and print to console
print("\nInitial Population:")
for idx, solution in enumerate(initial_population, 1):
    print(f"Solution {idx}: Variables = {solution['variables']}")

save_population_to_template(
    population=initial_population,
    template_file=template_file,
    output_file="Test_1.xlsx",
    sheet_name="simple",
    start_row=2,
    start_col=1
)

# Run SolidWorks and perform setup
dest_dir = r"D:\Close_loop_in_silico_optimization_showcase\T_1"
macro_file = copy_and_rename_macro_file(src_macro_file, dest_dir, 1)

changes = {
    r"Close_loop_in_silico_optimization_showcase\Design": r"Close_loop_in_silico_optimization_showcase\T_1\Design",
    r"Close_loop_in_silico_optimization_showcase\Test.xlsx": r"Close_loop_in_silico_optimization_showcase\Test_1.xlsx"
}

modified_content = update_bas_file(original_bas_file_path, changes, 1)
iteration_dir = os.path.join(output_directory, "T_1")
os.makedirs(iteration_dir, exist_ok=True)

output_file_name = "Creating3D_new.bas"
output_file_path = os.path.join(output_directory, output_file_name)

with open(output_file_path, 'w') as file:
    file.writelines(modified_content)

print(f"Modified .bas file saved at: {output_file_path}")

macro_file = r"D:\Close_loop_in_silico_optimization_showcase\T_1\test_T_1.swp"
open_sldprt_and_run_macro(
    file_path,
    macro_file,
    macro_module_name,
    macro_procedure_name,
    macro_module_name2,
    macro_procedure_name2,
)

# Run STAR-CCM+
input_java_file = r"D:\Close_loop_in_silico_optimization_showcase\Run_CFD.java"
output_java_file = r"C:\Program Files\Siemens\17.04.008\STAR-CCM+17.04.008\star\bin\Run_CFD_Modified.java"
folder_path = r"D:\Close_loop_in_silico_optimization_showcase\T_1"

old_string = "T_0"
new_string = "T_1"

replace_strings_and_update_population(input_java_file, output_java_file, old_string, new_string, folder_path)
run_starccm(output_java_file)

output_folder = os.path.join(dest_dir, 'output')
summary_file = os.path.join(output_folder, 'summary.csv')
process_all_csv_files(dest_dir, output_folder, summary_file)


# Real-time plotting setup
plt.ion()
fig, ax = plt.subplots()
ax.set_xlim(0, 0.6)  # Mixing Index range
ax.set_ylim(0, 16)  # Pressure Drop range
ax.set_xlabel('Mixing Index')
ax.set_ylabel('Pressure Drop')
plt.title('Autonmous in-silico optimization')
plt.grid(True)


# Adjust the size and position of the window
manager = plt.get_current_fig_manager()
manager.window.setGeometry(1400, 150, 800, 500)  # x, y, width, height
manager.window.raise_()  # Raise the window to the front
manager.window.activateWindow()  # Focus the window

# Colormap for dynamic colors
colors = plt.cm.get_cmap('tab10')

# Store scatter plots for each generation
scatter_plots = {}
legend_labels = []

# Initial plot setup for Generation 1
mixing_indices, pressure_drops = read_summary_csv(summary_file)

scatter_plots[f"Generation {1}"] = ax.scatter(
    mixing_indices, pressure_drops, label="Initial population", color=colors(1 % 10)
)

# Update the legend dynamically
ax.legend(loc='upper right')

# Redraw the plot
plt.draw()
plt.pause(0.1)  # Allow GUI event processing


# Optimization loop
for i in range(2, generations + 2):
    print(f"\n--- Generation {i} ---")

    # Perform non-dominated sorting and calculate metrics
    fronts = non_dominated_sorting(initial_population)
    pareto_front = [(sol["objectives"][0], sol["objectives"][1]) for sol in fronts[0]]
    hv = calculate_hypervolume(pareto_front, reference_point)

    # Display Pareto front and metrics
    print(f"\nPareto front at generation {i}:")
    for sol in fronts[0]:
        print(f"Variables = {sol['variables']}, Objectives = {sol['objectives']}")
    print(f"HyperVolume: {hv:.4f}")

    # Generate offspring
    offspring = []
    while len(offspring) < population_size:
        parent1 = tournament_selection(initial_population)
        parent2 = tournament_selection(initial_population)
        child1, child2 = crossover(parent1, parent2, crossover_rate, problem)
        off1 = mutate(child1, mutation_rate, problem)
        if not is_duplicate(off1, offspring):
            offspring.append(off1)
        if len(offspring) < population_size:
            off2 = mutate(child2, mutation_rate, problem)
            if not is_duplicate(off2, offspring):
                offspring.append(off2)

    # Print offspring to console
    print("\nGenerated Offspring Population:")
    for idx, solution in enumerate(offspring, 1):
        print(f"Offspring {idx}: Variables = {solution['variables']}")

    # Save offspring population to Test_{generation + 1}.xlsx in 'simple' sheet
    save_population_to_template(
        population=offspring,
        template_file=template_file,
        output_file=f"Test_{i}.xlsx",
        sheet_name="simple",
        start_row=2,
        start_col=1
    )

        
    ## Run solidWorks
    dest_dir = rf"D:\Close_loop_in_silico_optimization_showcase\T_{i}"
    macro_file = copy_and_rename_macro_file(src_macro_file, dest_dir, i)
    
    # Define changes for each iteration
    changes = {
        r"Close_loop_in_silico_optimization_showcase\Design": rf"Close_loop_in_silico_optimization_showcase\T_{i}\Design",
        r"Close_loop_in_silico_optimization_showcase\Test.xlsx": rf"Close_loop_in_silico_optimization_showcase\Test_{i}.xlsx"
        }

    # Call the function to get modified content
    modified_content = update_bas_file(original_bas_file_path, changes, i)

    # Create the output directory for this iteration
    iteration_dir = os.path.join(output_directory, f"T_{i}")
    os.makedirs(iteration_dir, exist_ok=True)

    # Save the modified content to a new file with the format `Creating3D_T_i.bas`
    output_file_name = "Creating3D_new.bas"
    output_file_path = os.path.join(output_directory, output_file_name)

    # Save the modified content
    with open(output_file_path, 'w') as file:
        file.writelines(modified_content)

    print(f"Modified .bas file saved at: {output_file_path}")

    # Create the location of macro file
    macro_file = rf"D:\Close_loop_in_silico_optimization_showcase\T_{i}\test_T_{i}.swp"

    # Run macro
    open_sldprt_and_run_macro(
        file_path,
        macro_file,
        macro_module_name,
        macro_procedure_name,
        macro_module_name2,
        macro_procedure_name2,
        )

        
    ## Run starccm+
    input_java_file = r"D:\Close_loop_in_silico_optimization_showcase\Run_CFD.java" 
    output_java_file = rf"C:\Program Files\Siemens\17.04.008\STAR-CCM+17.04.008\star\bin\Run_CFD_Modified.java"  
    folder_path = rf"D:\Close_loop_in_silico_optimization_showcase\T_{i}"  

    old_string = "T_0"  
    new_string = f"T_{i}"  

    # Update the file
    replace_strings_and_update_population(input_java_file, output_java_file, old_string, new_string, folder_path)
    
    # Open CMD and run starccm+ orders
    run_starccm(output_java_file)
    
    ## Data processing
    # Specify the input folder containing CSV files and the output folder for Excel files
    output_folder = os.path.join(dest_dir, 'output')
    summary_file = os.path.join(output_folder, 'summary.csv')
        
    # Process all CSV files in the input folder and create the summary
    process_all_csv_files(dest_dir, output_folder, summary_file)

    
    # Plotting updated population
    # Read data from the summary CSV for the current generation

    mixing_indices, pressure_drops = read_summary_csv(rf"D:\Close_loop_in_silico_optimization_showcase\T_{i}\output\summary.csv")

    # Add scatter plot for this generation
    scatter_plots[f"Generation {i}"] = ax.scatter(
        mixing_indices, pressure_drops, label=f"Population {i}", color=colors(i % 10)
    )
    
    # Update the legend dynamically
    ax.legend(loc='upper right')
    
    # Redraw the plot
    plt.draw()
    plt.pause(0.1)  # Allow GUI event processing
        
    
    ## Automatically read fitness values for the current generation
    if not os.path.exists(summary_file):
        raise FileNotFoundError(f"Fitness file '{summary_file}' not found in folder '{folder_path}'.")

    print(f"\nLoading fitness values for offspring from '{summary_file}'.")
    offspring = evaluate_offspring_from_file(offspring, summary_file)

    # Combine population and offspring
    initial_population += offspring

    # Perform non-dominated sorting and select the next generation
    fronts = non_dominated_sorting(initial_population)
    next_generation = []
    unique_solutions = set()  # Track unique solutions

    for front in fronts:
        for solution in front:
            solution_tuple = tuple(sorted(solution["variables"]))
            if solution_tuple not in unique_solutions:
                next_generation.append(solution)
                unique_solutions.add(solution_tuple)
            if len(next_generation) >= population_size:
                break
        if len(next_generation) >= population_size:
            break

    initial_population = next_generation

# Finalize plot
plt.ioff()  # Disable interactive mode
plt.show()  # Display the final plot




